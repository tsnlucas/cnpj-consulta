import requests
import openpyxl
import re
import time
from PyQt5.QtCore import QThread, pyqtSignal

class CNPJConsultaThread(QThread):
    # Sinal para atualizar a interface com informações sobre o progresso da consulta
    progress_signal = pyqtSignal(str)
    # Sinal para atualizar a barra de progresso
    progress_percent_signal = pyqtSignal(float)

    def __init__(self, arquivo_excel):
        super().__init__()
        self.arquivo_excel = arquivo_excel

    def run(self):
        workbook = openpyxl.load_workbook(self.arquivo_excel)
        sheet = workbook.active

        # Encontrar a coluna de CNPJ pelo rótulo
        coluna_cnpj = None
        for col in sheet.iter_cols(values_only=True):
            if 'CNPJ' in col:
                coluna_cnpj = col
                break

        if not coluna_cnpj:
            self.progress_signal.emit('Coluna de CNPJ não encontrada na planilha.')
            return

        total_cnpjs = len(coluna_cnpj) - 1  # Desconta o cabeçalho
        novo_nome_arquivo = self.arquivo_excel.replace('.xlsx', '_editado.xlsx')
        novo_workbook = openpyxl.Workbook()
        novo_sheet = novo_workbook.active

        novo_sheet['A1'] = 'CNPJ'
        novo_sheet['B1'] = 'Razão Social'
        novo_sheet['C1'] = 'Data de Abertura'
        novo_sheet['D1'] = 'Responsável 1'
        novo_sheet['E1'] = 'Responsável 2'
        novo_sheet['F1'] = 'Situação'

        linhas_processadas = 0
        erros = []

        for cnpj in coluna_cnpj[1:]:
            cnpj = str(cnpj).zfill(14)  # Preenche com zeros à esquerda para garantir 14 dígitos

            self.progress_signal.emit(f'Processando CNPJ: {cnpj}')

            max_tentativas = 5  # Aumente o número de tentativas se desejar
            tentativa = 1

            while tentativa <= max_tentativas:
                url = f'https://www.receitaws.com.br/v1/cnpj/{cnpj}'

                headers = {'Cache-Control': 'no-cache'}  # Adicione o cabeçalho de controle de cache

                response = requests.get(url, headers=headers)  # Use o cabeçalho na solicitação

                if response.status_code == 200:
                    dados_receita = response.json()
                    razao_social = dados_receita.get('nome', 'N/A')
                    abertura = dados_receita.get('abertura', 'N/A')
                    responsaveis = dados_receita.get('qsa', [{'nome': 'N/A'}])

                    # Inicialize nomes dos responsáveis com valores padrão
                    responsavel1 = 'N/A'
                    responsavel2 = 'N/A'

                    # Preencha os nomes dos responsáveis, se disponíveis
                    for i, resp in enumerate(responsaveis):
                        if i == 0:
                            responsavel1 = resp['nome']
                        elif i == 1:
                            responsavel2 = resp['nome']
                        else:
                            break

                    situacao = dados_receita.get('situacao', 'N/A')

                    novo_sheet.append([cnpj, razao_social, abertura, responsavel1, responsavel2, situacao])
                    linhas_processadas += 1
                    break  # Saia do loop while após obter os dados com sucesso

                else:
                    tentativa += 1

            # Calcular o progresso como uma porcentagem e emitir o sinal
            progress_percent = (linhas_processadas / total_cnpjs) * 100
            self.progress_percent_signal.emit(progress_percent)

            # Aguarde 30 segundos antes da próxima consulta
            if linhas_processadas < total_cnpjs:
                self.progress_signal.emit('Esperando 30 segundos antes da próxima consulta...')
                time.sleep(30)

        novo_workbook.save(novo_nome_arquivo)
        novo_workbook.close()
        workbook.close()

        if erros:
            erro_str = '\n'.join(erros)
            self.progress_signal.emit(f'Processo concluído com {linhas_processadas} de {total_cnpjs} CNPJs processados.\nErros:\n{erro_str}')
        else:
            self.progress_signal.emit(f'Processo concluído com {linhas_processadas} de {total_cnpjs} CNPJs processados.')
